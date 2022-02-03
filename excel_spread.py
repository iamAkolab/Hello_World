import openpyxl
import os

# change director
os.chdir(r'C:\\Users\\john\\Downloads')

# open a workbook
workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

# open sheet in workbook
sheet_one = workbook.get_sheet_by_name('Sheet1')
print(type(sheet_one))

# print sheet names
print(workbook.get_sheet_names())

# get cell
cellA1 = sheet_one['A1']
print(cellA1.value)

print(str(sheet_one['A1']))

# get cell
cellB1 = sheet_one['B1']
print(cellB1.value)

# get cell
cellC1 = sheet_one['C1']
print(cellC1.value)

# get cell
sheet_one.cell(row=1, column=2)

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)